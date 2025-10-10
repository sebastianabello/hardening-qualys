from __future__ import annotations
from pathlib import Path
import gzip, csv, sys
from typing import Dict, Iterator
import json
from elasticsearch import Elasticsearch, helpers
from openpyxl import load_workbook
from .settings import settings
from itertools import chain

try:
    csv.field_size_limit(10 * 1024 * 1024)
except Exception:
    csv.field_size_limit(min(sys.maxsize, 2_147_483_647))

def _iter_csv_docs(csv_path: Path):
    import json
    opener = gzip.open if csv_path.suffix.lower()==".gz" else open
    mode = "rt" if csv_path.suffix.lower()==".gz" else "r"
    with opener(csv_path, mode, encoding="utf-8", newline="", errors="ignore") as f:
        rdr = csv.reader(f)
        headers = next(rdr, None)
        if not headers:
            return
        headers = [str(h) if h is not None else "" for h in headers]
        for row in rdr:
            obj = {}
            for i, h in enumerate(headers):
                obj[h] = "" if i >= len(row) or row[i] is None else row[i]
            yield json.dumps(obj, ensure_ascii=False)


def _iter_excel_docs(xlsx_path: Path) -> Iterator[str]:
    """
    Itera filas de un XLSX (read-only) y devuelve cada fila como JSON (str).
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            return
        headers = [str(h) if h is not None else "" for h in headers]
        for values in rows:
            obj = {}
            for i, h in enumerate(headers):
                obj[h] = "" if values is None or i >= len(values) or values[i] is None else values[i]
            yield json.dumps(obj, ensure_ascii=False)
    finally:
        wb.close()


def _get_index_name_from_file(file_name: str) -> tuple[str, bool]:
    """
    Genera el nombre del índice basado en el nombre del archivo.
    Usa la misma lógica que _nombre_base() para mantener consistencia.
    """
    import re
    from pathlib import Path
    
    # Obtener el nombre base del archivo sin extensión
    base_name = Path(file_name).stem  # Elimina .xlsx, .csv, .gz
    
    # Si termina en .csv (para archivos .csv.gz), quitarlo también
    if base_name.endswith('.csv'):
        base_name = base_name[:-4]
    
    # Convertir el nombre del archivo a un índice válido para Elasticsearch
    # Mantener la estructura: cliente-hardening-[control-statics]-fecha[-ajustado]
    index_name = base_name.lower()
    
    # Reemplazar caracteres especiales con guiones bajos
    index_name = re.sub(r'[^a-z0-9\-]', '_', index_name)  # Mantener guiones
    index_name = re.sub(r'_+', '_', index_name)  # Múltiples _ → uno solo
    index_name = index_name.strip('_')  # Eliminar _ al inicio y final
    
    # Detectar si es una versión "ajustada" (igual que en tu código)
    ajustada = "ajustado" in file_name.lower()
    
    print(f"📁 Archivo: {file_name} → Índice: {index_name} (ajustada: {ajustada})")
    
    return index_name, ajustada


def _create_elasticsearch_client() -> Elasticsearch:
    """
    Crea el cliente de Elasticsearch usando API Key únicamente.
    """
    if not settings.ES_BASE_URL:
        raise ValueError("ES_BASE_URL no está configurado. Configura la URL de Elasticsearch en el archivo .env")
    
    if not settings.ES_API_KEY:
        raise ValueError("ES_API_KEY no está configurado. Configura el API Key en el archivo .env")
    
    print("🔑 Conectando con API Key")
    
    # Configuración igual a tu elastic_uploader.py
    es_config = {
        "hosts": [settings.ES_BASE_URL],
        "api_key": settings.ES_API_KEY,
        "verify_certs": settings.ES_VERIFY_CERTS,
        "request_timeout": 120
    }
    
    return Elasticsearch(**es_config)


async def ingest_run_folder(run_output_dir: Path) -> Dict[str, int]:
    """
    Recorre los XLSX del run y los ingesta en ES vía bulk API.
    Devuelve conteo por índice.
    """
    count_by_index: Dict[str, int] = {}
    
    # Crear cliente de Elasticsearch
    es = _create_elasticsearch_client()
    
    try:
        # Verificar conexión
        info = es.info()
        print(f"✅ Conectado a Elasticsearch: {info['cluster_name']} (v{info['version']['number']})")
        
        files = list(chain(sorted(run_output_dir.glob("*.xlsx")),
                           sorted(run_output_dir.glob("*.csv.gz")),
                           sorted(run_output_dir.glob("*.csv"))))

        for file_path in files:
            print(f"Procesando archivo: {file_path.name}")
            index, ajustada = _get_index_name_from_file(file_path.name)
            print(f"Destino: índice={index}, ajustada={ajustada}")

            # Preparar documentos para bulk insert
            def generate_docs():
                if file_path.suffix.lower() in [".csv", ".gz"]:
                    iter_docs = _iter_csv_docs(file_path)
                else:
                    iter_docs = _iter_excel_docs(file_path)
                
                for doc_line in iter_docs:
                    doc = json.loads(doc_line)
                    if "ajustada" not in doc:
                        doc["ajustada"] = ajustada
                    
                    yield {
                        "_index": index,
                        "_source": doc
                    }

            # Usar helpers.bulk para inserción eficiente
            try:
                success_count, failed_items = helpers.bulk(
                    es,
                    generate_docs(),
                    chunk_size=1000,
                    max_chunk_bytes=5_000_000,
                    request_timeout=120,
                    raise_on_error=False
                )
                
                print(f"✅ {success_count} documentos indexados en '{index}'")
                if failed_items:
                    print(f"⚠️ {len(failed_items)} documentos fallaron")
                    # Log primeros errores para debug
                    for error in failed_items[:3]:
                        print(f"   Error: {error}")
                
                count_by_index[index] = count_by_index.get(index, 0) + success_count
                
            except Exception as e:
                print(f"❌ Error procesando archivo {file_path.name}: {e}")
                raise

    finally:
        es.close()

    return count_by_index
